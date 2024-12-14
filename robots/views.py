import json
import datetime
from openpyxl import Workbook
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Robot
from django.utils.dateparse import parse_datetime
from django.db.models import Count

@csrf_exempt
def create_robot(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            serial = data['serial']
            model = data['model']
            version = data['version']
            created = parse_datetime(data['created'])

            if not model or not version or not created:
                return JsonResponse({'Ошибка': 'Не верные данные.'}, status=400)

            existing_robot = Robot.objects.filter(model=model, version=version).first()

            if existing_robot:
                serial = existing_robot.serial
            else:
                return JsonResponse({'Ошибка': f'Робота модели {model} и версии {version} не найдены в базе.'}, status=400)

            robot = Robot.objects.create(serial = serial, model=model, version=version, created=created, in_stock=True)
            robot.save()
            return JsonResponse({'Сообщение': 'Запись создана успешно'}, status=201)

        except (KeyError, ValueError, json.JSONDecodeError) as e:
            return JsonResponse({'Ошибка': str(e)}, status=400)

    else:
        return JsonResponse({'Ошибка': 'Доступен только POST-method'}, status=405)


def generate_excel_report(request):
    end_date = datetime.datetime.now()
    start_date = end_date - datetime.timedelta(days=7)

    robots = Robot.objects.filter(created__range=(start_date, end_date))

    wb = Workbook()

    robot_models = robots.values('model').distinct()

    sheet_created = False

    for robot_model in robot_models:
        model_name = robot_model['model']
        ws = wb.create_sheet(title=model_name)
        ws.append(['Модель', 'Версия', 'Количество за неделю'])

        model_robots = robots.filter(model=model_name)
        versions = model_robots.values('version').annotate(total=Count('id'))

        for version in versions:
            ws.append([model_name, version['version'], version['total']])
        
        sheet_created = True

    if 'Sheet' in wb.sheetnames and sheet_created:
        del wb['Sheet']

    if not sheet_created:
        ws = wb.active
        ws.title = "No Data"
        ws.append(['Нет данных за последнюю неделю'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="robots_report.xlsx"'
    wb.save(response)
    return response